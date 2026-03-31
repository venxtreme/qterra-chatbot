from fastapi import FastAPI, APIRouter, HTTPException
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
import google.generativeai as genai
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from typing import List, Optional
import os
import json
from datetime import datetime
import openpyxl

import properties_db


def load_leased_properties(filepath: str = "Leased_properties_with_URLs.xlsx") -> list:
    """Load previously leased properties from the Excel file."""
    leased = []
    try:
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(row):
                leased.append({
                    "status": row[0] or "",
                    "address": row[1] or "",
                    "property_type": row[2] or "",
                    "beds": row[3] or "",
                    "baths": row[4] or "",
                    "sqft": row[5] or "",
                    "price": row[6] or "",
                    "url": row[7] or "",
                })
    except Exception as e:
        print(f"WARNING: Could not load leased properties: {e}")
    return leased


LEASED_PROPERTIES = load_leased_properties()
print(f"Loaded {len(LEASED_PROPERTIES)} leased properties for owner assurance.")

def load_env():
    """Load .env file for local development. Skipped silently in production."""
    try:
        with open(".env", "r") as f:
            for line in f:
                if "=" in line and not line.startswith("#"):
                    key, value = line.strip().split("=", 1)
                    os.environ.setdefault(key, value)
    except FileNotFoundError:
        pass  # Running in production with env vars set directly

load_env()

app = FastAPI()
api_router = APIRouter()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

api_key = os.environ.get("GEMINI_API_KEY")
if not api_key:
    raise RuntimeError("GEMINI_API_KEY is not set in .env")
genai.configure(api_key=api_key)

scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']

# Support credentials from file (local) or env var (production/Railway)
google_creds_json = os.environ.get("GOOGLE_CREDENTIALS_JSON")
if google_creds_json:
    import tempfile
    creds_data = json.loads(google_creds_json)
    with tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False) as tmp:
        json.dump(creds_data, tmp)
        tmp_path = tmp.name
    creds = ServiceAccountCredentials.from_json_keyfile_name(tmp_path, scope)
else:
    creds = ServiceAccountCredentials.from_json_keyfile_name('google_credentials.json', scope)

client = gspread.authorize(creds)
spreadsheet_id = os.environ.get("SPREADSHEET_ID")
sheet = None

try:
    sheet = client.open_by_key(spreadsheet_id).sheet1
    print("Successfully connected to Google Sheet")
except Exception as e:
    print(f"WARNING: Could not connect to Google Sheet. Error: {e}")


class Message(BaseModel):
    role: str
    content: str

class ChatRequest(BaseModel):
    messages: List[Message]


SYSTEM_PROMPT = """
Your name is Quinn. You are a warm, empathetic, and professional chatbot for Qterra Property Management in Ontario, Canada.
You have a soft, human personality — you show genuine care, use friendly affirmations like "That's great!", "Of course!", "Absolutely!", "I completely understand!", and never sound robotic.
Keep every response SHORT and conversational. Ask only ONE question at a time. Never rush the user.

============================
STEP 1 — GREETING & INTENT
============================
Start by warmly greeting the user and asking how you can help.
Listen carefully to their opening message and determine their role:

- TENANT: they mention looking for a place to rent, a condo, basement, apartment, house, etc.
- OWNER: they mention having a property, looking for tenants, wanting to rent out, etc.
- PROPERTY MANAGEMENT: they mention property management services, managing their property, etc.

If you are not sure, gently ask: "Are you looking to rent a place, or are you a property owner?"

============================
STEP 2 — COLLECT NAME & PHONE (ALL ROLES)
============================
Once you know the role, collect:
1. First ask for their Name.
2. Then ask for their Phone Number.

PHONE NUMBER VALIDATION RULES (apply strictly):
- Valid Canadian/US format: NPA-NXX-XXXX
  - NPA = Area code: first digit must be 2-9 (NOT 0 or 1)
  - NXX = Exchange: first digit must be 2-9 (NOT 0 or 1)
  - Remaining digits (XXXX): any digit 0-9
- Dashes are optional — accept both 6475559919 and 647-555-9919
- REJECT numbers where area code or exchange starts with 0 or 1 (e.g. 123-456-7890 or 011-555-1234 are INVALID)
- If invalid, kindly say something like: "Hmm, that doesn't look like a valid phone number. Could you double-check and share it again? It should be a 10-digit number like 647-555-9919."
- Only proceed once you have a valid phone number.

============================
STEP 3A — TENANT FLOW
============================
After Name + Phone, collect the following ONE AT A TIME:
1. Preferred location or area in Ontario.
2. Type of property (Condo, Basement, House, Townhouse, etc.)
3. Preferred move-in date.
4. Credit score range (e.g. 600-650, 700-750). Say something warm like: "To help us find the best fit for you, could you share your approximate credit score range? This is just to help us pre-screen options for you."
5. Number of occupants who will be living in the unit.

AFTER collecting all 5 items:
- Recommend up to 3 matching available properties from the database (provided in context). Always include the full URL link for each.
- Let them know they can apply here: https://forms.zohopublic.com/quettapropertymanagement/form/RentalApplication/formperma/-nWZTD2qFkCIqpQG-9edv2W5AHgXpfi8DUFlR_k7SNg
- Thank them warmly.

============================
STEP 3B — OWNER FLOW
============================
After Name + Phone, collect ONE AT A TIME:
1. Location/address of their property.
2. Type of property.
3. When they'd like tenants to move in (move-in date / availability date).

After collecting all info:
- Warmly assure them: "Thank you so much! I've passed your information to our team and someone will be in touch with you very soon."
- Add reassurance: "Just so you know, we have successfully placed tenants in many similar properties across Ontario. You're in great hands!"
- Thank them and wish them well.

============================
STEP 3C — PROPERTY MANAGEMENT FLOW
============================
After Name + Phone, ask:
1. Location of their property or area they need help with.

Then say warmly: "Wonderful! I've noted your information and our property management team will reach out to you as soon as possible. We're excited to help you!"
Thank them.

============================
FINAL STEP — JSON PAYLOAD (INTERNAL, DO NOT SHOW TO USER)
============================
Once you have all the required information for the role, you MUST end your FINAL response with a JSON block (after your goodbye message) formatted EXACTLY like this:

```json
{
  "name": "Jane Smith",
  "role": "Tenant",
  "location": "Brampton",
  "property_type": "Condo",
  "phone": "647-555-9919",
  "move_in_date": "May 1, 2026",
  "credit_score": "700-750",
  "num_occupants": "2",
  "summary": "Tenant looking for a condo in Brampton, moving in May 2026, credit score 700-750, 2 occupants."
}
```

For Owners:
```json
{
  "name": "John Doe",
  "role": "Owner",
  "location": "Mississauga",
  "property_type": "Detached House",
  "phone": "905-444-1234",
  "move_in_date": "June 1, 2026",
  "credit_score": "",
  "num_occupants": "",
  "summary": "Owner has a detached house in Mississauga available June 2026."
}
```

For Property Management:
```json
{
  "name": "Ali Khan",
  "role": "Property Management",
  "location": "Ottawa",
  "property_type": "",
  "phone": "613-999-8888",
  "move_in_date": "",
  "credit_score": "",
  "num_occupants": "",
  "summary": "Interested in property management services in Ottawa."
}
```

IMPORTANT RULES:
- Do NOT use markdown formatting like ** or bullet points in your chat messages — keep it plain, natural, conversational.
- Do NOT ask all questions at once. ONE question per response.
- Always be warm, patient, and encouraging.
- The JSON block must always appear at the very end of your message and only once.
"""

model = genai.GenerativeModel('gemini-2.5-flash', system_instruction=SYSTEM_PROMPT)


def determine_properties(messages_content: str):
    """Extract location and property type from conversation to query the database."""
    location = None
    property_type = None

    cities = [
        "hamilton", "scarborough", "toronto", "whitby", "brockville", "mississauga",
        "richmond hill", "milton", "nepean", "innisfil", "orleans", "etobicoke",
        "brampton", "oshawa", "ajax", "pickering", "barrie", "ottawa", "london",
        "kitchener", "waterloo", "cambridge", "guelph", "burlington", "oakville",
        "markham", "vaughan", "newmarket", "north york", "york", "concord",
        "woodbridge", "welland", "niagara falls", "st. catharines", "kingston",
        "stoney creek", "waterdown", "east york", "ridgetown", "king", "mono",
        "east gwillimbury", "innisfil beach", "whitchurch-stouffville"
    ]
    types = [
        "condo", "basement", "house", "detached", "townhouse", "semi-detached",
        "apartment", "multi-plex", "upper level", "main floor", "stacked townhouse"
    ]

    lower_content = messages_content.lower()

    for c in cities:
        if c in lower_content:
            location = c
            break

    for t in types:
        if t in lower_content:
            property_type = t
            break

    return properties_db.search_properties(location=location, property_type=property_type), location, property_type


@api_router.post("/chat")
async def chat_endpoint(req: ChatRequest):
    try:
        history = []
        full_conversation = ""
        for m in req.messages[:-1]:
            history.append({"role": "model" if m.role == "assistant" else "user", "parts": [m.content]})
            full_conversation += f"{m.role}: {m.content}\n"

        chat = model.start_chat(history=history)

        last_msg = req.messages[-1].content
        full_conversation += f"user: {last_msg}\n"

        # Detect conversation type
        lower_convo = full_conversation.lower()
        is_owner_convo = any(word in lower_convo for word in [
            "owner", "my property", "looking for tenant", "rent out", "have a property",
            "i own", "landlord", "list my", "find tenants"
        ])
        is_tenant_convo = not is_owner_convo and any(word in lower_convo for word in [
            "tenant", "looking for", "rent", "find a", "need a place", "move in", "looking to rent",
            "i need", "searching for", "apartment", "condo", "basement", "townhouse"
        ])

        context = ""

        # Inject matching available properties for Tenants
        if is_tenant_convo:
            properties, location, property_type = determine_properties(full_conversation)
            if properties and location and property_type:
                context = "\n[SYSTEM: The user is a Tenant. Here are up to 3 matching available properties. Include each property's full URL link in your message:\n"
                for p in properties:
                    price = f" | ${p.get('price', 'N/A')}/mo" if p.get('price') else ""
                    beds = f" | {p.get('beds', '')} bed" if p.get('beds') else ""
                    baths = f" | {p.get('baths', '')} bath" if p.get('baths') else ""
                    link = f" | URL: {p.get('url', '')}"
                    context += f"- {p['type']} at {p['address']}{price}{beds}{baths}{link}\n"
                context += "Always output the exact URLs so the user can click them.]\n"

        # Inject matching leased properties for Owners as reassurance
        if is_owner_convo and LEASED_PROPERTIES:
            _, location, property_type = determine_properties(full_conversation)
            matches = []
            for lp in LEASED_PROPERTIES:
                loc_match = location and location.lower() in lp["address"].lower()
                type_match = property_type and property_type.lower() in lp["property_type"].lower()
                if loc_match or type_match:
                    matches.append(lp)
                if len(matches) >= 3:
                    break
            # Fall back to first 3 if no specific match
            if not matches:
                matches = LEASED_PROPERTIES[:3]
            context += "\n[SYSTEM: The user is an Owner. After collecting their info and assuring them the team will follow up, "
            context += "show these examples of similar properties we have successfully leased to build their confidence. Include the URLs:\n"
            for lp in matches:
                context += f"- {lp['property_type']} at {lp['address']} | {lp['price']}/mo | URL: {lp['url']}\n"
            context += "]\n"

        response = chat.send_message(last_msg + context)
        response_text = response.text

        # Parse and save JSON payload if present
        if "```json" in response_text:
            json_str = response_text.split("```json")[1].split("```")[0].strip()
            try:
                data = json.loads(json_str)
                if sheet:
                    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    row = [
                        now,
                        data.get("name", ""),
                        data.get("role", ""),
                        data.get("location", ""),
                        data.get("property_type", ""),
                        data.get("phone", ""),
                        data.get("move_in_date", ""),
                        data.get("credit_score", ""),
                        data.get("num_occupants", ""),
                        data.get("summary", ""),
                    ]
                    sheet.append_row(row)
                else:
                    print(f"Lead captured (Sheets not connected): {data}")
            except Exception as e:
                print(f"Failed to process JSON or save to Sheets: {e}")

            # Strip JSON from response shown to user
            response_text = response_text.split("```json")[0].strip()

        return {"response": response_text}

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


app.include_router(api_router, prefix="/api")
app.mount("/", StaticFiles(directory="static", html=True), name="static")

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
